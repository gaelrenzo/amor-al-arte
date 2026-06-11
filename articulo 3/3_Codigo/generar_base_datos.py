import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

np.random.seed(42)
n = 120

t = pd.date_range(start='2015-01-01', periods=n, freq='ME')

trend = np.linspace(0, 5, n)
cycle = 0.5 * np.sin(np.linspace(0, 8 * np.pi, n))
shock_PIB = np.random.normal(0, 0.3, n)
PIB = 100 + trend + cycle + shock_PIB + 0.3 * np.cumsum(np.random.normal(0, 0.1, n))

shock_INFL = np.random.normal(0, 0.2, n)
INFL = 3.5 + 0.15 * np.cumsum(shock_INFL) + 0.1 * cycle

shock_TC = np.random.normal(0, 0.4, n)
TC = 3.8 + 0.2 * np.cumsum(shock_TC) + 0.05 * (PIB - 100)

shock_TI = np.random.normal(0, 0.25, n)
TI = 5.0 + 1.5 * (INFL - 3.5) + 0.3 * np.cumsum(shock_TI)

shock_M2 = np.random.normal(0, 0.3, n)
M2 = np.log(500 + 10 * trend + np.cumsum(shock_M2))

shock_G = np.random.normal(0, 0.2, n)
G = 20 + 0.15 * trend + np.cumsum(shock_G)

base = pd.DataFrame({
    'Fecha': t,
    'PIB': np.round(PIB, 4),
    'INFLACION': np.round(INFL, 4),
    'TIPO_CAMBIO': np.round(TC, 4),
    'TASA_INTERES': np.round(TI, 4),
    'OFERTA_MONETARIA': np.round(M2, 4),
    'GASTO_GOBIERNO': np.round(G, 4),
})

wb = Workbook()
for hoja, datos, color in [
    ('Datos Originales', base, '1F4E79'),
    ('PIB', base[['Fecha', 'PIB']], '2E75B6'),
    ('Inflacion', base[['Fecha', 'INFLACION']], 'C00000'),
    ('TipoCambio', base[['Fecha', 'TIPO_CAMBIO']], '00B050'),
    ('TasaInteres', base[['Fecha', 'TASA_INTERES']], 'ED7D31'),
    ('OfertaMonetaria', base[['Fecha', 'OFERTA_MONETARIA']], '7030A0'),
    ('GastoGobierno', base[['Fecha', 'GASTO_GOBIERNO']], '4472C4'),
]:
    if hoja == 'Datos Originales':
        ws = wb.active
        ws.title = hoja
    else:
        ws = wb.create_sheet(title=hoja)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    
    for col_idx, col_name in enumerate(datos.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    data_font = Font(name='Calibri', size=11)
    for row_idx, (_, row) in enumerate(datos.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 1:
                cell.number_format = 'YYYY-MM-DD'
            else:
                cell.number_format = '#,##0.0000'
    
    ws.column_dimensions['A'].width = 15
    for i in range(2, len(datos.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 17
    ws.freeze_panes = 'A2'

ws_meta = wb.create_sheet(title='Metadatos', index=0)
meta_info = [
    ('METADATOS - BASE DE DATOS MACROECONÓMICA SINTÉTICA', ''),
    ('', ''),
    ('Curso', 'Econometría III'),
    ('Universidad', 'Universidad Nacional del Altiplano'),
    ('Facultad', 'Facultad de Ingeniería Económica'),
    ('Modelo', 'VAR Estructural (SVAR)'),
    ('Periodo', '2015-01 a 2024-12'),
    ('Frecuencia', 'Mensual'),
    ('Observaciones', str(n)),
    ('Propósito', 'Fines académicos y aprendizaje metodológico'),
    ('', ''),
    ('VARIABLES:', ''),
    ('PIB', 'Producto Bruto Interno (índice base 100)'),
    ('INFLACION', 'Tasa de inflación (%)'),
    ('TIPO_CAMBIO', 'Tipo de cambio nominal (S/./USD)'),
    ('TASA_INTERES', 'Tasa de interés de referencia (%)'),
    ('OFERTA_MONETARIA', 'Logaritmo de la oferta monetaria'),
    ('GASTO_GOBIERNO', 'Gasto público (índice)'),
    ('', ''),
    ('RESTRICCIONES DE IDENTIFICACIÓN (ORDEN CHOLESKY):', ''),
    ('1. Gasto Gobierno', 'No responde contemporáneamente a shocks de otras variables'),
    ('2. Oferta Monetaria', 'Responde a Gasto Gobierno, pero no a PIB, inflación, TC ni TI'),
    ('3. PIB', 'Responde a política fiscal y monetaria, pero no a inflación, TC ni TI'),
    ('4. Inflación', 'Responde a política fiscal, monetaria y PIB, pero no a TC ni TI'),
    ('5. Tipo de Cambio', 'Responde a todas excepto a Tasa de Interés'),
    ('6. Tasa de Interés', 'Responde contemporáneamente a todas las variables'),
]

for row_idx, (key, val) in enumerate(meta_info, 1):
    cell_k = ws_meta.cell(row=row_idx, column=1, value=key)
    cell_v = ws_meta.cell(row=row_idx, column=2, value=val)
    if row_idx == 1:
        cell_k.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    elif ':' in key and val == '':
        cell_k.font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    else:
        cell_k.font = Font(name='Calibri', size=11)
        cell_v.font = Font(name='Calibri', size=11)
    ws_meta.column_dimensions['A'].width = 35
    ws_meta.column_dimensions['B'].width = 80

output = r'C:\Users\renzo\amor-al-arte\articulo 3\base_datos_sintetica.xlsx'
wb.save(output)
print(f'Base de datos generada: {output}')
